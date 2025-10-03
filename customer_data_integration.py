#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
顧客データ統合スクリプト
各顧客フォルダ内の「顧客ファイル.xlsx」を統合し、一つのExcelファイルに全シートを保存する
"""

import os
import re
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import logging
from pathlib import Path

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('customer_integration.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class CustomerDataIntegrator:
    def __init__(self, customer_folder_path, output_path=None):
        """
        顧客データ統合クラス
        
        Args:
            customer_folder_path (str): 顧客フォルダのパス
            output_path (str): 出力ファイルのパス（Noneの場合は自動生成）
        """
        self.customer_folder_path = Path(customer_folder_path)
        self.output_path = output_path or self._generate_output_path()
        self.processed_count = 0
        self.error_count = 0
        self.error_log = []
        
    def _generate_output_path(self):
        """出力ファイルパスを生成"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"顧客データ統合_全シート_{timestamp}.xlsx"
    
    def _extract_customer_info(self, folder_name):
        """
        フォルダ名から顧客情報を抽出
        
        Args:
            folder_name (str): フォルダ名（例: "0001あきずみ秋住幸信様(ｾﾘｶ)"）
            
        Returns:
            tuple: (顧客番号, 顧客名, 車種名)
        """
        # 顧客番号を抽出（先頭の4桁）
        customer_number = folder_name[:4]
        
        # 顧客名を抽出（番号の後から「様」まで）
        name_match = re.search(r'(\d{4})(.+?)様', folder_name)
        if name_match:
            customer_name = name_match.group(2)
        else:
            customer_name = folder_name[4:]
        
        # 車種名を抽出（括弧内）
        car_match = re.search(r'\((.+?)\)', folder_name)
        car_name = car_match.group(1) if car_match else ""
        
        return customer_number, customer_name, car_name
    
    def _generate_sheet_name(self, customer_number, customer_name, car_name, sheet_name):
        """
        シート名を生成（Excelの制限に合わせて31文字以内）
        
        Args:
            customer_number (str): 顧客番号
            customer_name (str): 顧客名
            car_name (str): 車種名
            sheet_name (str): 元のシート名
            
        Returns:
            str: 生成されたシート名
        """
        # 基本形式: 0001_顧客名_車種名_シート名
        base_name = f"{customer_number}_{customer_name}_{car_name}_{sheet_name}"
        
        # 31文字制限に合わせて調整
        if len(base_name) > 31:
            # 顧客名を短縮
            max_name_length = 31 - len(f"{customer_number}__{car_name}_{sheet_name}")
            if max_name_length > 0:
                customer_name = customer_name[:max_name_length]
                base_name = f"{customer_number}_{customer_name}_{car_name}_{sheet_name}"
            else:
                # それでも長い場合は番号とシート名のみ
                base_name = f"{customer_number}_{sheet_name}"
        
        return base_name
    
    def _copy_sheet(self, source_wb, source_sheet_name, target_wb, target_sheet_name):
        """
        シートをコピー
        
        Args:
            source_wb: ソースワークブック
            source_sheet_name (str): ソースシート名
            target_wb: ターゲットワークブック
            target_sheet_name (str): ターゲットシート名
        """
        try:
            # ソースシートを取得
            source_sheet = source_wb[source_sheet_name]
            
            # ターゲットシートを作成
            target_sheet = target_wb.create_sheet(title=target_sheet_name)
            
            # セルの値をコピー
            for row in source_sheet.iter_rows():
                for cell in row:
                    target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                    target_cell.value = cell.value
                    
                    # スタイルもコピー（安全な方法）
                    try:
                        if cell.has_style:
                            if cell.font:
                                target_cell.font = openpyxl.styles.Font(
                                    name=cell.font.name,
                                    size=cell.font.size,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic,
                                    vertAlign=cell.font.vertAlign,
                                    underline=cell.font.underline,
                                    strike=cell.font.strike,
                                    color=cell.font.color
                                )
                            if cell.border:
                                target_cell.border = openpyxl.styles.Border(
                                    left=cell.border.left,
                                    right=cell.border.right,
                                    top=cell.border.top,
                                    bottom=cell.border.bottom
                                )
                            if cell.fill:
                                target_cell.fill = openpyxl.styles.PatternFill(
                                    fill_type=cell.fill.fill_type,
                                    start_color=cell.fill.start_color,
                                    end_color=cell.fill.end_color
                                )
                            if cell.number_format:
                                target_cell.number_format = cell.number_format
                            if cell.alignment:
                                target_cell.alignment = openpyxl.styles.Alignment(
                                    horizontal=cell.alignment.horizontal,
                                    vertical=cell.alignment.vertical,
                                    text_rotation=cell.alignment.text_rotation,
                                    wrap_text=cell.alignment.wrap_text,
                                    shrink_to_fit=cell.alignment.shrink_to_fit,
                                    indent=cell.alignment.indent
                                )
                    except Exception as style_error:
                        # スタイルコピーでエラーが発生しても処理を続行
                        logger.warning(f"スタイルコピーでエラー（無視）: {style_error}")
            
            # 列幅をコピー
            for col in source_sheet.column_dimensions:
                if source_sheet.column_dimensions[col].width:
                    target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
            
            # 行の高さをコピー
            for row in source_sheet.row_dimensions:
                if source_sheet.row_dimensions[row].height:
                    target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
            
            logger.info(f"シート '{source_sheet_name}' を '{target_sheet_name}' としてコピーしました")
            
        except Exception as e:
            logger.error(f"シートコピーエラー: {e}")
            raise
    
    def _create_index_sheet(self, wb, customer_list):
        """
        目次シートを作成
        
        Args:
            wb: ワークブック
            customer_list (list): 顧客情報のリスト
        """
        index_sheet = wb.create_sheet(title="目次", index=0)
        
        # ヘッダー行
        headers = ["顧客番号", "顧客名", "車種名", "シート名", "ファイルパス"]
        for col, header in enumerate(headers, 1):
            index_sheet.cell(row=1, column=col, value=header)
        
        # データ行
        row = 2
        for customer_info in customer_list:
            customer_number, customer_name, car_name, sheet_names, file_path = customer_info
            for sheet_name in sheet_names:
                index_sheet.cell(row=row, column=1, value=customer_number)
                index_sheet.cell(row=row, column=2, value=customer_name)
                index_sheet.cell(row=row, column=3, value=car_name)
                index_sheet.cell(row=row, column=4, value=sheet_name)
                index_sheet.cell(row=row, column=5, value=file_path)
                row += 1
        
        # 列幅を調整
        for col in range(1, 6):
            index_sheet.column_dimensions[chr(64 + col)].width = 20
    
    def _create_log_sheet(self, wb):
        """
        処理ログシートを作成
        
        Args:
            wb: ワークブック
        """
        log_sheet = wb.create_sheet(title="処理ログ")
        
        # ヘッダー行
        headers = ["処理日時", "処理件数", "エラー件数", "エラー内容"]
        for col, header in enumerate(headers, 1):
            log_sheet.cell(row=1, column=col, value=header)
        
        # 統計情報
        log_sheet.cell(row=2, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        log_sheet.cell(row=2, column=2, value=self.processed_count)
        log_sheet.cell(row=2, column=3, value=self.error_count)
        
        # エラーログ
        row = 4
        log_sheet.cell(row=row, column=1, value="エラー詳細:")
        row += 1
        
        for error in self.error_log:
            log_sheet.cell(row=row, column=1, value=error)
            row += 1
        
        # 列幅を調整
        for col in range(1, 5):
            log_sheet.column_dimensions[chr(64 + col)].width = 30
    
    def process_customer_folders(self):
        """
        顧客フォルダを処理して統合Excelファイルを作成
        """
        logger.info("顧客データ統合処理を開始します")
        logger.info(f"顧客フォルダパス: {self.customer_folder_path}")
        logger.info(f"出力ファイル: {self.output_path}")
        
        # 出力用ワークブックを作成
        output_wb = Workbook()
        # デフォルトシートを削除
        output_wb.remove(output_wb.active)
        
        customer_list = []
        
        try:
            # 顧客フォルダを取得（番号順でソート）
            customer_folders = sorted([f for f in self.customer_folder_path.iterdir() 
                                     if f.is_dir() and f.name.startswith(('0', '1', '2', '3'))])
            
            logger.info(f"処理対象フォルダ数: {len(customer_folders)}")
            
            for folder_path in customer_folders:
                try:
                    self._process_single_customer(folder_path, output_wb, customer_list)
                    self.processed_count += 1
                    
                    # 進捗表示
                    if self.processed_count % 10 == 0:
                        logger.info(f"処理済み: {self.processed_count}件")
                        
                except Exception as e:
                    self.error_count += 1
                    error_msg = f"フォルダ '{folder_path.name}' の処理でエラー: {str(e)}"
                    logger.error(error_msg)
                    self.error_log.append(error_msg)
            
            # 目次シートとログシートを作成
            self._create_index_sheet(output_wb, customer_list)
            self._create_log_sheet(output_wb)
            
            # ファイルを保存
            output_wb.save(self.output_path)
            
            logger.info(f"統合処理が完了しました")
            logger.info(f"処理件数: {self.processed_count}")
            logger.info(f"エラー件数: {self.error_count}")
            logger.info(f"出力ファイル: {self.output_path}")
            
        except Exception as e:
            logger.error(f"統合処理でエラーが発生しました: {str(e)}")
            raise
        finally:
            output_wb.close()
    
    def _process_single_customer(self, folder_path, output_wb, customer_list):
        """
        単一顧客フォルダを処理
        
        Args:
            folder_path (Path): 顧客フォルダのパス
            output_wb: 出力ワークブック
            customer_list (list): 顧客情報リスト
        """
        customer_file_path = folder_path / "顧客ファイル.xlsx"
        
        if not customer_file_path.exists():
            raise FileNotFoundError(f"顧客ファイルが見つかりません: {customer_file_path}")
        
        # 顧客情報を抽出
        customer_number, customer_name, car_name = self._extract_customer_info(folder_path.name)
        
        # 顧客ファイルを開く
        try:
            source_wb = openpyxl.load_workbook(customer_file_path, data_only=True)
            sheet_names = source_wb.sheetnames
            
            # 各シートをコピー
            copied_sheets = []
            for sheet_name in sheet_names:
                target_sheet_name = self._generate_sheet_name(
                    customer_number, customer_name, car_name, sheet_name
                )
                
                # シート名の重複チェック
                original_name = target_sheet_name
                counter = 1
                while target_sheet_name in output_wb.sheetnames:
                    target_sheet_name = f"{original_name}_{counter}"
                    counter += 1
                
                self._copy_sheet(source_wb, sheet_name, output_wb, target_sheet_name)
                copied_sheets.append(target_sheet_name)
            
            # 顧客情報をリストに追加
            customer_list.append((
                customer_number, customer_name, car_name, 
                copied_sheets, str(customer_file_path)
            ))
            
            source_wb.close()
            
        except Exception as e:
            raise Exception(f"ファイル読み込みエラー: {str(e)}")

def main():
    """メイン処理"""
    # 顧客フォルダのパス
    customer_folder_path = "☆顧客フォルダ"
    
    # 統合処理を実行
    integrator = CustomerDataIntegrator(customer_folder_path)
    integrator.process_customer_folders()

if __name__ == "__main__":
    main()
